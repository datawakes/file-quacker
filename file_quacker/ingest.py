"""File ingestion → DuckDB tables / views.

Text files load as views over `read_csv(..., all_varchar=True, ...)` so
source values are preserved exactly; no silent date / int coercion.
`escape` mirrors `quote` (DuckDB's CSV convention) so RFC-4180 doubled-
quote escapes parse correctly.  All identifiers pass through
`db.quote_ident` to prevent SQL injection via table / column names.

For text files we try a ladder of strategies and keep the first one that
loads cleanly.  DuckDB's built-in `encoding=` covers the big three
(UTF-8, UTF-16, Latin-1) plus `cp1252` for Windows text, without having
to transcode to a temp file.  `chardet` seeds the ladder with a
best-guess encoding; the final lenient strategy flips
`ignore_errors=True` so malformed rows land in a `reject_errors` table
instead of aborting the load.
"""

from __future__ import annotations

import csv as _csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal

import chardet
import duckdb

from . import db

CandidateDelim = Literal[',', '|', ';', '\t']
CANDIDATE_DELIMS: tuple[CandidateDelim, ...] = (',', '|', ';', '\t')

NO_QUOTE_SENTINEL = '«'

SQL_SAFE_NAME = re.compile(r'[^0-9A-Za-z_]+')

# Metadata column prepended to every loaded table/view.  Survives any
# downstream filter or sort, so the operator can always recover a row's
# position in the source file.
ROW_NUM_COLUMN = '_src_row_num'


@dataclass
class IngestOptions:
    delimiter: str | None = None
    quote: str = '"'
    escape: str | None = None
    encoding: str | None = None
    first_row: int = 1
    skip_rows: int = 0
    ignore_errors: bool = False
    # parallel=False is required when null_padding meets quoted newlines.
    parallel: bool = True
    # strict_mode=False relaxes RFC-4180 quoting; the catch-all rung uses it.
    strict_mode: bool = True
    # -1 scans the whole file for inference; None uses DuckDB's 20480 default.
    # Lifts column-count detection past late wide rows.
    sample_size: int | None = None
    # Views lazy-read the source on every query and tank interactive perf;
    # always materialize.
    materialize: bool = True
    sheet: str | None = None
    preferred_table_name: str | None = None
    # Case-sensitive (DuckDB binary collation), so the surname `Null`,
    # lowercase `na`, country code `Na` etc. survive.  Empty string is
    # intentionally absent: DuckDB already nulls unquoted empty fields
    # in CSV, and a blanket `'' -> NULL` rule loses too much in parquet/xlsx.
    nullify_tokens: list[str] = field(
        default_factory=lambda: ['NA', 'NaN']
    )
    # Frontend sets this after the "load this one or all in folder?"
    # prompt; > 1 entry triggers `read_parquet([...])`.
    parquet_paths: list[str] | None = None
    # Excel preamble cropping.  `excel_header_row` is 1-based (so 3 means
    # "skip two title rows"); `excel_start_col`/`end_col` are A-style.
    excel_header_row: int | None = None
    excel_start_col: str | None = None
    excel_end_col: str | None = None


@dataclass
class IngestResult:
    name: str
    source_path: str
    format: Literal['csv', 'parquet', 'xlsx']
    materialized: bool
    row_count: int
    col_count: int
    has_rejects: bool
    columns: list[dict] = field(default_factory=list)
    # Effective ingest options; what DuckDB actually parsed with.  For
    # non-text formats the delimiter/encoding fields are None.
    delimiter: str | None = None
    encoding: str | None = None
    first_row: int | None = None
    # Name of the strategy that succeeded (for file-format debugging).
    strategy: str | None = None


# --------------------------------------------------------------------------- #
# Public entry points                                                         #
# --------------------------------------------------------------------------- #

def preview_file_lines(path: str, n: int = 30, encoding: str | None = None) -> dict:
    """First ``n`` raw lines of a text file plus the encoding used.
    Feeds the Open-with-options dialog.  Decode failures fall back to
    UTF-8 + replace so the user always sees something (even mojibake)
    and can pick a different encoding from the dropdown.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)
    if not encoding:
        encoding = _chardet_to_duckdb(_detect_encoding(p)) or 'utf-8'
    used = encoding
    try:
        with p.open('r', encoding=encoding, errors='replace', newline='') as f:
            lines = []
            for i, line in enumerate(f):
                if i >= n:
                    break
                lines.append(line.rstrip('\n').rstrip('\r'))
    except (LookupError, UnicodeDecodeError):
        with p.open('rb') as f:
            raw = f.read()
        lines = raw.decode('utf-8', errors='replace').splitlines()[:n]
        used = 'utf-8 (forced)'
    return {'lines': lines, 'encoding': used, 'sniffed_delimiter': _sniff_delimiter_bytes(p)}


def load_file(path: str, opts: IngestOptions | None = None) -> IngestResult:
    opts = opts or IngestOptions()
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)
    ext = p.suffix.lower()
    if ext == '.parquet':
        return _load_parquet(p, opts)
    # .xls / .xlsx / .xlsm extensions sometimes lie; ELT pipelines
    # commonly transform Excel files to pipe-delimited text and keep
    # the original extension so downstream tooling stays unchanged.
    # Sniff the first 8 bytes; if it's not the Excel magic, fall
    # through to the CSV ladder.
    if ext in ('.xlsx', '.xlsm', '.xls') and not _looks_like_excel(p):
        return _load_csv(p, opts)
    if ext == '.xlsx':
        return _load_xlsx(p, opts)
    return _load_csv(p, opts)


# --------------------------------------------------------------------------- #
# CSV / delimited text                                                        #
# --------------------------------------------------------------------------- #

def _load_csv(path: Path, opts: IngestOptions) -> IngestResult:
    """Load a delimited-text file by trying a ladder of parsing
    strategies.  The first one that produces a usable table wins; each
    failed attempt drops its partial artifact.  If the user passed an
    explicit encoding / delimiter we honor that and skip the ladder.
    """
    # Delimiter sniff is encoding-agnostic; we count bytes of the four
    # candidate separators, all of which are ASCII, so we don't have to
    # successfully decode the file first.
    delim = opts.delimiter or _sniff_delimiter_bytes(path)

    # User-forced encoding bypasses the ladder.
    if opts.encoding:
        return _try_load_csv(path, opts, delim, encoding=opts.encoding,
                             ignore_errors=opts.ignore_errors,
                             strategy='user-set')

    guessed = _detect_encoding(path)
    guessed_duck = _chardet_to_duckdb(guessed)
    # Encoding for the lenient/fallback rungs.  If chardet landed on
    # something other than utf-8/ascii, trust it and carry it forward;
    # otherwise cp1252 is the safe every-byte-decodable guess.  This
    # avoids silently re-reading a utf-16 or latin-1 file as cp1252 when
    # the original encoding rung failed for a non-encoding reason.
    lenient_enc = guessed_duck if (guessed_duck and guessed_duck != 'utf-8') else 'cp1252'

    # (label, encoding, ignore_errors, no_quote, parallel, strict_mode, sample_size, auto_delim)
    strategies: list[tuple[str, str, bool, bool, bool, bool, int | None, bool]] = []
    strategies.append(('utf-8', 'utf-8', False, False, True, True, None, False))
    # Rescues clean files when byte-count sniff picks the wrong delim
    # (e.g. a pipe file with commas in quoted narratives).
    strategies.append(('utf-8 + auto-delim', 'utf-8', False, False, True, True, None, True))
    if guessed_duck and guessed_duck != 'utf-8':
        strategies.append((f'{guessed_duck} (chardet)', guessed_duck, False, False, True, True, None, False))
    # Try UTF-8 again with sequential + full-sample BEFORE switching to
    # lenient_enc; otherwise a real UTF-8 file with quoted newlines or
    # a late wide row gets silently re-decoded as cp1252.
    strategies.append(('utf-8 + sequential + full-sample',
                       'utf-8', False, False, False, True, -1, False))
    if lenient_enc != 'utf-8':
        strategies.append((f'{lenient_enc} + sequential + full-sample',
                           lenient_enc, False, False, False, True, -1, False))
    strategies.append((f'{lenient_enc} + sequential + ignore_errors + full-sample',
                       lenient_enc, True, False, False, True, -1, False))
    strategies.append((f'{lenient_enc} + sequential + ignore_errors + auto-delim + full-sample',
                       lenient_enc, True, False, False, True, -1, True))
    # Last resort: drop the quote char and loosen RFC-4180 strictness.
    strategies.append((f'{lenient_enc} + sequential + ignore_errors + no-quote + strict=false',
                       lenient_enc, True, True, False, False, -1, False))

    errors: list[str] = []
    for label, enc, lenient, no_quote, parallel, strict, sample_size, auto_delim in strategies:
        try:
            effective = _with_flags(opts, no_quote=no_quote, parallel=parallel,
                                    strict=strict, sample_size=sample_size)
            effective_delim = None if auto_delim else delim
            return _try_load_csv(path, effective, effective_delim,
                                 encoding=enc,
                                 ignore_errors=lenient,
                                 strategy=label)
        except _RETRY_ERRORS as ex:
            # Keep all of DuckDB's error lines joined with `|`; the first
            # line is the type, the rest name the column / value / context.
            detail = ' | '.join(line.strip() for line in str(ex).splitlines() if line.strip())
            errors.append(f'[{label}] {detail}')
            continue

    raise IngestError(
        'Could not read this file after trying '
        f'{len(strategies)} strategies:\n  ' + '\n  '.join(errors)
    )


def _with_flags(
    opts: IngestOptions,
    *,
    no_quote: bool,
    parallel: bool,
    strict: bool,
    sample_size: int | None,
) -> IngestOptions:
    from dataclasses import replace
    kwargs: dict = {'parallel': parallel, 'strict_mode': strict, 'sample_size': sample_size}
    if no_quote:
        kwargs['quote'] = NO_QUOTE_SENTINEL
        kwargs['escape'] = None
    return replace(opts, **kwargs)


def _try_load_csv(path: Path, opts: IngestOptions, delim: str | None, *,
                  encoding: str, ignore_errors: bool,
                  strategy: str) -> IngestResult:
    """Execute one load attempt.  Any DuckDB error bubbles up; the
    caller (ladder loop) decides whether to retry.  Pass ``delim=None``
    to let DuckDB sniff the delimiter itself."""
    quote = '' if opts.quote == NO_QUOTE_SENTINEL else opts.quote
    escape = quote if opts.escape is None else opts.escape

    header = opts.first_row >= 1
    skip = opts.skip_rows + (opts.first_row - 1 if opts.first_row > 1 else 0)

    name = opts.preferred_table_name or _derive_table_name(path)
    name = _unique_name(name)
    qname = db.quote_ident(name)

    object_kw = 'table' if opts.materialize else 'view'

    params: dict = {
        'header': header,
        'quote': quote,
        'escape': escape,
        'ignore_errors': ignore_errors,
        'all_varchar': True,
        'null_padding': True,
        'skip': skip,
        'encoding': encoding,
        'parallel': opts.parallel,
        'strict_mode': opts.strict_mode,
        # Default is 2 MB; insufficient for exports that pack long
        # narratives into a single cell.  32 MB covers pathological
        # rows without blowing up memory on sane files.
        'max_line_size': 32 * 1024 * 1024,
    }
    if opts.sample_size is not None:
        params['sample_size'] = opts.sample_size
    if delim is not None:
        params['delim'] = delim
    if opts.nullify_tokens:
        # Passing nullstr= replaces DuckDB's default (empty string for
        # unquoted empty fields), so prepend '' to keep that behavior
        # while adding our extra tokens.
        params['nullstr'] = ['', *opts.nullify_tokens]
    if ignore_errors:
        params['store_rejects'] = True

    sql = f"""
        create or replace {object_kw} {qname} as
        select  row_number() over ()  as {db.quote_ident(ROW_NUM_COLUMN)}
        ,       *
        from    read_csv(
                    {_sql_path_literal(str(path))}
                  , {_fmt_params(params)}
                )
        ;
    """
    c = db.conn()
    c.execute(sql)

    if opts.first_row == 0:
        _rename_default_columns(name)

    row_count, col_count, cols = _describe(name)

    has_rejects = False
    if ignore_errors:
        rejects_name = _unique_name(f'{name}__rejects')
        try:
            c.execute(f"""
                create or replace table {db.quote_ident(rejects_name)} as
                select  *
                from    reject_errors
                ;
            """)
            c.execute("""
                drop table if exists reject_errors
                ;
            """)
            has_rejects = True
        except duckdb_error_types():
            has_rejects = False

    return IngestResult(
        name=name,
        source_path=str(path),
        format='csv',
        materialized=opts.materialize,
        row_count=row_count,
        col_count=col_count,
        has_rejects=has_rejects,
        columns=cols,
        delimiter=delim,
        encoding=encoding,
        first_row=opts.first_row,
        strategy=strategy,
    )


class IngestError(Exception):
    """Raised when every strategy in the ladder fails."""


# Error classes we retry on: anything DuckDB raises for "file not
# parseable as configured"; invalid encoding, column-count mismatch,
# conversion errors.  Programmer errors (SyntaxError, AttributeError)
# never get caught.
_RETRY_ERRORS: tuple[type, ...] = (
    duckdb.InvalidInputException,
    duckdb.ConversionException,
    duckdb.Error,
)


# --------------------------------------------------------------------------- #
# Parquet                                                                     #
# --------------------------------------------------------------------------- #

def _load_parquet(path: Path, opts: IngestOptions) -> IngestResult:
    paths = opts.parquet_paths or [str(path)]
    if opts.preferred_table_name:
        name = opts.preferred_table_name
    elif len(paths) > 1:
        # Multi-file load; name the table after the parent directory
        # (the dataset's "name" in Spark/Hive output conventions).
        # Reuse `_derive_table_name` so the dir name goes through the
        # same snake_case + sql-safe sanitizer as the file-stem path.
        parent_dir = Path(paths[0]).parent
        name = _derive_table_name(parent_dir) if parent_dir.name else _derive_table_name(path)
    else:
        name = _derive_table_name(path)
    name = _unique_name(name)
    qname = db.quote_ident(name)
    object_kw = 'table' if opts.materialize else 'view'
    if len(paths) == 1:
        from_clause = f'read_parquet({_sql_path_literal(paths[0])})'
        source_path = paths[0]
    else:
        # `union_by_name=true` lets sibling parquets with slightly
        # different column orderings (or extra/missing columns) union
        # cleanly; missing columns become NULL.
        list_sql = '[' + ', '.join(_sql_path_literal(p) for p in paths) + ']'
        from_clause = f'read_parquet({list_sql}, union_by_name=true)'
        source_path = f'{Path(paths[0]).parent} ({len(paths)} parquet files)'
    sql = f"""
        create or replace {object_kw} {qname} as
        select  row_number() over ()  as {db.quote_ident(ROW_NUM_COLUMN)}
        ,       *
        from    {from_clause}
        ;
    """
    db.conn().execute(sql)
    if opts.materialize and opts.nullify_tokens:
        _nullify_string_tokens(name, opts.nullify_tokens)
    row_count, col_count, cols = _describe(name)
    return IngestResult(
        name=name, source_path=source_path, format='parquet',
        materialized=opts.materialize, row_count=row_count,
        col_count=col_count, has_rejects=False, columns=cols,
    )


def inspect_parquet_set(path: str) -> dict:
    """List sibling .parquet files in the same directory.

    Returns ``{this_path, peers, total_bytes}`` where ``peers`` always
    includes the original path.  The frontend uses this to ask
    "load just this file or all parquet files in the folder?" before
    deciding which paths to pass to ``_load_parquet``.
    """
    p = Path(path)
    if not p.exists() or p.suffix.lower() != '.parquet':
        size = p.stat().st_size if p.exists() else 0
        return {'this_path': str(p), 'peers': [str(p)], 'total_bytes': size}

    parent = p.parent
    # Case-insensitive match; Windows is fine with .PARQUET / .parquet
    # naming variants in the same directory.
    seen: set[str] = set()
    peers: list[Path] = []
    for entry in sorted(parent.iterdir()):
        if not entry.is_file():
            continue
        if entry.suffix.lower() != '.parquet':
            continue
        key = str(entry).lower()
        if key in seen:
            continue
        seen.add(key)
        peers.append(entry)
    if not peers:
        peers = [p]

    total_bytes = sum(pp.stat().st_size for pp in peers)
    return {
        'this_path': str(p),
        'peers': [str(pp) for pp in peers],
        'total_bytes': int(total_bytes),
    }


def _nullify_string_tokens(table_name: str, tokens: list[str]) -> None:
    """Set every VARCHAR cell whose value matches one of the placeholder
    tokens (NA, N/A, NaN, …) to NULL.  Only meaningful for materialized
    tables (a view would have nothing to update)."""
    if not tokens:
        return
    c = db.conn()
    qname = db.quote_ident(table_name)
    cols = c.execute(f"""
        pragma table_info({qname})
        ;
    """).fetchall()
    for row in cols:
        col_name = row[1]
        col_type = row[2].upper()
        if col_type != 'VARCHAR':
            continue
        if col_name == ROW_NUM_COLUMN:
            continue
        qcol = db.quote_ident(col_name)
        placeholders = ', '.join(['?'] * len(tokens))
        c.execute(f"""
            update  {qname}
            set     {qcol} = null
            where   {qcol} in ({placeholders})
            ;
        """, list(tokens))


# --------------------------------------------------------------------------- #
# Excel                                                                       #
# --------------------------------------------------------------------------- #

def _load_xlsx(path: Path, opts: IngestOptions) -> IngestResult:
    c = db.conn()
    c.execute("""
        install excel
        ;
    """)
    c.execute("""
        load excel
        ;
    """)

    name = opts.preferred_table_name or _derive_table_name(path)
    if opts.sheet and not opts.preferred_table_name:
        # Disambiguate when loading multiple sheets from the same file.
        name = f'{name}_{_derive_table_name(Path(opts.sheet))}'
    name = _unique_name(name)
    qname = db.quote_ident(name)
    object_kw = 'table' if opts.materialize else 'view'

    params: dict = {
        'all_varchar': True,
        'empty_as_varchar': True,
        'ignore_errors': True,
    }
    if opts.sheet:
        params['sheet'] = opts.sheet
    # Preamble cropping.  When the auto-detect step (or the user)
    # tells us the data doesn't start at A1, pass a `range=` with a
    # computed end column and `stop_at_empty=true` so trailing empty
    # rows aren't materialized into millions of NULL rows.  When
    # there's no preamble we let DuckDB pick the bounds itself.
    start_col = (opts.excel_start_col or 'A').upper()
    header_row = max(1, opts.excel_header_row or 1)
    end_col = (opts.excel_end_col or '').upper() or None
    params['header'] = True
    if header_row > 1 or start_col != 'A' or end_col is not None:
        # `range='B3:XFD'` would have DuckDB allocate ~16k columns
        # and OOM.  When the caller didn't pin an end column, probe
        # it via openpyxl so we always pass a bounded range.
        if end_col is None:
            end_col = _detect_end_col_on_row(path, opts.sheet, header_row, start_col)
        params['range'] = f'{start_col}{header_row}:{end_col}'
        params['stop_at_empty'] = True

    def _build_sql(p: Path, kw: str) -> str:
        return f"""
            create or replace {kw} {qname} as
            select  row_number() over ()  as {db.quote_ident(ROW_NUM_COLUMN)}
            ,       *
            from    read_xlsx(
                        {_sql_path_literal(str(p))}
                      , {_fmt_params(params)}
                    )
            ;
        """

    cleaned: Path | None = None
    try:
        c.execute(_build_sql(path, object_kw))
    except duckdb.Error as ex:
        # DuckDB's excel reader is stricter than Excel about styles.xml
        # / xf entries; files exported by some tools (Aspose, certain
        # SaaS dashboards) parse fine in Excel but error here.  Re-save
        # the workbook via openpyxl, which rewrites styles.xml from
        # scratch, and retry once.  The retry forces TABLE mode (even
        # if the caller asked for a view) so we can safely delete the
        # cleaned temp file after load; a view would have nothing to
        # read once the temp is gone.
        if not _is_styles_xml_error(ex):
            raise
        try:
            cleaned = _resave_xlsx_for_duckdb(path)
        except Exception:
            # Couldn't even open in openpyxl; surface DuckDB's
            # original message rather than the secondary failure.
            raise ex from None
        try:
            c.execute(_build_sql(cleaned, 'table'))
        except duckdb.Error:
            cleaned.unlink(missing_ok=True)
            raise
    finally:
        if cleaned is not None and cleaned.exists():
            cleaned.unlink(missing_ok=True)

    if opts.materialize and opts.nullify_tokens:
        _nullify_string_tokens(name, opts.nullify_tokens)
    row_count, col_count, cols = _describe(name)
    return IngestResult(
        name=name, source_path=str(path), format='xlsx',
        materialized=opts.materialize, row_count=row_count,
        col_count=col_count, has_rejects=False, columns=cols,
    )


def _is_styles_xml_error(ex: BaseException) -> bool:
    """True for DuckDB excel-reader complaints about styles.xml / xf
    entries.  These are the symptoms of cosmetic XLSX malformations
    that Excel tolerates but the bundled excel extension rejects."""
    msg = str(ex).lower()
    return 'styles.xml' in msg or 'xf entry' in msg or 'invalid xf' in msg


def _resave_xlsx_for_duckdb(path: Path) -> Path:
    """Open `path` with openpyxl and save to a sibling temp file with
    a clean styles.xml.  The returned path is the caller's to delete.
    Uses ``data_only=False`` so cell formulas survive the round-trip
    (DuckDB's reader will still see only the cached value)."""
    import tempfile
    import openpyxl
    wb = openpyxl.load_workbook(filename=str(path), read_only=False, data_only=False)
    fd, tmp = tempfile.mkstemp(prefix='fq_clean_', suffix='.xlsx',
                               dir=str(path.parent))
    import os
    os.close(fd)
    wb.save(tmp)
    wb.close()
    return Path(tmp)


def inspect_excel_file(path: str) -> dict:
    """Probe an Excel file for the chooser modal.

    Returns:
      ``{'kind': 'text', 'this_path': ...}`` when the file's bytes
      don't match the Excel magic; i.e. an ELT-transformed
      pipe-delimited file masquerading under an .xls/.xlsx extension.
      The frontend should fall back to the CSV ingest path.

      ``{'kind': 'excel', 'this_path': ..., 'sheets': [...]}``
      otherwise, with one entry per sheet:
      ``{'name': str, 'header_row': int, 'start_col': str,
         'detected': bool, 'preview': [[cell, ...], ...]}``.
      `header_row` is 1-based and `start_col` is a column letter; both
      are auto-detected from the densest non-empty row in the top-
      left corner of the sheet.  `detected` is True when the
      auto-detector chose anything other than A1; the frontend can
      use it to highlight that we cropped a preamble.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)
    if not _looks_like_excel(p):
        return {'kind': 'text', 'this_path': str(p), 'sheets': []}

    # openpyxl is already a project dependency; used in the export
    # writer.  Read-only mode is fast and memory-light.
    import openpyxl
    wb = openpyxl.load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        sheets: list[dict] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            preview, header_row, start_col_idx, end_col_idx = _excel_sheet_preview(ws)
            start_col = _col_letter(start_col_idx)
            end_col = _col_letter(end_col_idx)
            sheets.append({
                'name': sheet_name,
                'header_row': header_row,
                'start_col': start_col,
                'end_col': end_col,
                'detected': header_row > 1 or start_col != 'A',
                'preview': preview,
            })
        return {'kind': 'excel', 'this_path': str(p), 'sheets': sheets}
    finally:
        wb.close()


def load_excel_workbook(path: str, sheet_specs: list[dict],
                        opts: IngestOptions | None = None) -> list[dict]:
    """Load one or more sheets from an Excel file.  Each entry in
    ``sheet_specs`` is ``{'name', 'header_row', 'start_col'}``.  Returns
    a list of ``IngestResult``-as-dict.  The frontend gets back the
    summary for every sheet it asked for so it can list them in a
    success toast and refresh the sidebar."""
    base_opts = opts or IngestOptions()
    results: list[dict] = []
    for spec in sheet_specs:
        sub = IngestOptions(
            **{k: v for k, v in base_opts.__dict__.items()
               if k not in ('sheet', 'excel_header_row', 'excel_start_col',
                            'excel_end_col', 'preferred_table_name')},
            sheet=spec.get('name'),
            excel_header_row=int(spec.get('header_row') or 1),
            excel_start_col=(spec.get('start_col') or 'A'),
            excel_end_col=(spec.get('end_col') or None),
        )
        r = load_file(path, sub)
        from dataclasses import asdict
        results.append(asdict(r))
    return results


def _detect_end_col_on_row(path: Path, sheet: str | None,
                           header_row: int, start_col: str) -> str:
    """Open the workbook with openpyxl and return the rightmost
    non-empty column letter on ``header_row`` (1-based).  Used as a
    bounded fallback for DuckDB's `range=` argument so we never
    allocate the full 16k-column block.

    Falls back to ``start_col`` when the row turns out to be empty;
    that gives a single-column range which is benign and will hand
    back zero rows via `stop_at_empty=true`."""
    import openpyxl
    start_idx = _col_index_from_letter(start_col)
    wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        # Scan a generous block; openpyxl iter_rows is lazy so this
        # is cheap even for very wide sheets.
        rightmost = start_idx
        for row in ws.iter_rows(min_row=header_row, max_row=header_row,
                                min_col=1, max_col=16384, values_only=True):
            for j, v in enumerate(row, start=1):
                if v is not None and str(v) != '':
                    if j > rightmost:
                        rightmost = j
            break  # only one row
        return _col_letter(rightmost)
    finally:
        wb.close()


def _col_index_from_letter(letter: str) -> int:
    """Excel column letter → 1-based index ('A' → 1, 'AA' → 27)."""
    n = 0
    for ch in (letter or 'A').upper():
        if not ('A' <= ch <= 'Z'):
            continue
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n or 1


def _col_letter(idx: int) -> str:
    """1-based column index → Excel column letter (A, B, ..., Z, AA, AB, ...)."""
    if idx < 1:
        idx = 1
    out = ''
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = chr(ord('A') + rem) + out
    return out


def _excel_sheet_preview(ws, max_rows: int = 30, max_cols: int = 30) -> tuple[list[list[str]], int, int, int]:
    """Read a sample of the top-left of ``ws`` and return:
    ``(preview_grid, header_row, start_col, end_col)`` (all 1-based).

    Auto-detect strategy:
      * Walk the sample looking for the row with the highest density
        of non-empty cells.  That row is almost always the header in
        files with a preamble (title rows are sparse, data rows are
        dense, header rows are dense too).  Earliest densest row
        wins so we keep as many rows as possible.
      * The start / end columns are the leftmost / rightmost non-
        empty cells on that same row.
      * If nothing looks data-shaped, default to (1, 1, 1).
    """
    preview: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows,
                            min_col=1, max_col=max_cols,
                            values_only=True):
        preview.append(['' if v is None else str(v) for v in row])
    if not preview:
        return preview, 1, 1, 1

    best_density = 0
    best_row = 1
    best_left = 1
    best_right = 1
    for r_idx, row in enumerate(preview, start=1):
        nonempty_cols = [i for i, v in enumerate(row, start=1) if v != '']
        density = len(nonempty_cols)
        if density > best_density:
            best_density = density
            best_row = r_idx
            best_left = nonempty_cols[0]
            best_right = nonempty_cols[-1]

    # Sparse / non-tabular; don't pretend we found a header.
    if best_density < 2:
        return preview, 1, 1, 1
    return preview, best_row, best_left, best_right


# --------------------------------------------------------------------------- #
# Helpers                                                                     #
# --------------------------------------------------------------------------- #

def _looks_like_excel(path: Path) -> bool:
    """Magic-byte check.  `.xlsx` / `.xlsm` are ZIP archives starting
    with `PK\\x03\\x04`; `.xls` is an OLE compound document starting
    with `\\xd0\\xcf\\x11\\xe0\\xa1\\xb1\\x1a\\xe1`.  Anything else is a
    text file with a misleading extension."""
    try:
        with path.open('rb') as f:
            head = f.read(8)
    except OSError:
        return False
    return head.startswith(b'PK\x03\x04') or head.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1')


def _detect_encoding(path: Path, sample_bytes: int = 64 * 1024) -> str:
    with path.open('rb') as f:
        sample = f.read(sample_bytes)
    guess = chardet.detect(sample) or {}
    enc = (guess.get('encoding') or 'utf-8').lower()
    if enc in ('ascii',):
        return 'utf-8'
    return enc


# Chardet labels → DuckDB encoding names.  Chardet reports Python's
# codec names; DuckDB's CSV reader accepts a different (ICU-derived)
# set.  We map the four encodings that actually turn up in practice
# for this app's workload; UTF-8, UTF-16, Latin-1, and cp1252 (the
# catch-all for Windows-sourced text files).  Anything else chardet
# reports falls through to the ladder's cp1252 fallback; users with
# unusual encodings can type one into the ingest options.
_CHARDET_TO_DUCK: dict[str, str] = {
    'utf-8': 'utf-8',
    'utf-8-sig': 'utf-8',
    'ascii': 'utf-8',

    'utf-16': 'utf-16',
    'utf-16le': 'utf-16',
    'utf-16be': 'utf-16',

    'iso-8859-1': 'latin-1',
    'latin-1':    'latin-1',
    'iso-8859-15': 'latin-1',  # only the Euro sign differs

    'windows-1252': 'cp1252',
    'cp1252':       'cp1252',
}


def _chardet_to_duckdb(name: str) -> str | None:
    return _CHARDET_TO_DUCK.get(name.lower())


def _sniff_delimiter_bytes(path: Path, sample_bytes: int = 64 * 1024) -> str:
    """Pick the likeliest delimiter by counting bytes in a raw read.
    All candidate delimiters (``, | ; \\t``) are ASCII, so we don't
    have to decode the file first; this runs even when the file's
    encoding is unknown or wrong.
    """
    with path.open('rb') as f:
        sample = f.read(sample_bytes)
    try:
        text = sample.decode('utf-8', errors='replace')
        dialect = _csv.Sniffer().sniff(text, delimiters=''.join(CANDIDATE_DELIMS))
        if dialect.delimiter in CANDIDATE_DELIMS:
            return dialect.delimiter
    except _csv.Error:
        pass
    # Fall back: the candidate with the most byte hits wins.
    counts = {d: sample.count(d.encode('ascii')) for d in CANDIDATE_DELIMS}
    return max(CANDIDATE_DELIMS, key=lambda d: counts[d])


def _derive_table_name(path: Path) -> str:
    raw = path.stem
    clean = SQL_SAFE_NAME.sub('_', raw).strip('_').lower() or 'table'
    if clean[0].isdigit():
        clean = 't_' + clean
    return clean


def _unique_name(base: str) -> str:
    c = db.conn()
    rows = c.execute("""
        select  table_name
        from    information_schema.tables
        where   table_schema = current_schema()
        ;
    """).fetchall()
    existing = {r[0] for r in rows}
    if base not in existing:
        return base
    i = 2
    while f'{base}_{i}' in existing:
        i += 1
    return f'{base}_{i}'


def _rename_default_columns(name: str) -> None:
    """DuckDB names headerless columns column0, column1, ...  → col_0, col_1."""
    c = db.conn()
    rows = c.execute(f"""
        pragma table_info({db.quote_ident(name)})
        ;
    """).fetchall()
    for row in rows:
        old = row[1]
        if old.startswith('column') and old[len('column'):].isdigit():
            new = 'col_' + old[len('column'):]
            c.execute(f"""
                alter table {db.quote_ident(name)}
                rename column {db.quote_ident(old)} to {db.quote_ident(new)}
                ;
            """)


def _describe(name: str) -> tuple[int, int, list[dict]]:
    c = db.conn()
    rows = c.execute(f"""
        pragma table_info({db.quote_ident(name)})
        ;
    """).fetchall()
    cols = [{'name': r[1], 'type': r[2]} for r in rows]
    (rc,) = c.execute(f"""
        select  count(*)
        from    {db.quote_ident(name)}
        ;
    """).fetchone()
    return int(rc), len(cols), cols


def _sql_str_literal(s: str) -> str:
    """SQL string literal — single-quote escape only."""
    return "'" + s.replace("'", "''") + "'"


def _sql_path_literal(p: str) -> str:
    """SQL string literal for a filesystem path: also normalizes `\\`
    to `/` so Windows paths round-trip through DuckDB's read_csv /
    read_parquet without backslash-escape headaches."""
    return _sql_str_literal(p.replace('\\', '/'))


def _fmt_params(params: dict) -> str:
    """Format a dict as `key1 = val1, key2 = val2` for DuckDB function args."""
    parts = []
    for k, v in params.items():
        if isinstance(v, bool):
            parts.append(f"{k} = {'true' if v else 'false'}")
        elif isinstance(v, int):
            parts.append(f"{k} = {v}")
        elif isinstance(v, str):
            parts.append(f"{k} = {_sql_str_literal(v)}")
        elif isinstance(v, (list, tuple)):
            # DuckDB list literal: ['a', 'b'].  Used for read_csv args
            # like `nullstr` that accept multiple sentinel strings.
            inner = ', '.join(_sql_str_literal(str(x)) for x in v)
            parts.append(f"{k} = [{inner}]")
        else:
            raise TypeError(f'unsupported param type for {k}: {type(v).__name__}')
    return ', '.join(parts)


def duckdb_error_types() -> tuple[type, ...]:
    import duckdb
    return (duckdb.Error,)
